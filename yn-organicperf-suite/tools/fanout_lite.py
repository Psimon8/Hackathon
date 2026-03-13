"""
Fan-Out Lite — Application Tkinter standalone.
Génère les 5 queries fan-out les plus pertinentes par mot-clé via OpenAI GPT-4o-mini.
Langues supportées : FR, EN, DE, ES, PT-BR.
Usage : python tools/fanout_lite.py  (depuis la racine du projet)
"""

import os
import sys
import threading
import tkinter as tk
from concurrent.futures import ThreadPoolExecutor, as_completed
from tkinter import ttk, messagebox, filedialog
from typing import List

# ── Ensure project root is on sys.path so we can import core/modules ────────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from modules.fanout.generator import FanoutGenerator

# ── Language map ────────────────────────────────────────────────────────────
LANGUAGES = {
    "FR - Français": "fr",
    "EN - English": "en",
    "DE - Deutsch": "de",
    "ES - Español": "es",
    "PT-BR - Português": "pt",
}


# ═══════════════════════════════════════════════════════════════════════════
# Tkinter Application
# ═══════════════════════════════════════════════════════════════════════════

class FanoutLiteApp:
    """Lightweight Tkinter GUI for Fan-Out generation."""

    # ── Colors / Style ──────────────────────────────────────────────────
    BG = "#1e1e2e"
    FG = "#cdd6f4"
    ACCENT = "#89b4fa"
    ACCENT_HOVER = "#74c7ec"
    SURFACE = "#313244"
    SURFACE2 = "#45475a"
    GREEN = "#a6e3a1"
    RED = "#f38ba8"
    YELLOW = "#f9e2af"
    FONT = ("Segoe UI", 10)
    FONT_BOLD = ("Segoe UI", 10, "bold")
    FONT_TITLE = ("Segoe UI", 16, "bold")
    FONT_MONO = ("Consolas", 10)

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("🌐 Fan-Out Lite Generator")
        self.root.geometry("960x720")
        self.root.minsize(800, 600)
        self.root.configure(bg=self.BG)

        # State
        self._results: List[dict] = []  # [{keyword, topic, queries: [(rank, query)], error}]
        self._generating = False

        self._apply_style()
        self._build_ui()
        self._load_env_key()

    # ── Style ───────────────────────────────────────────────────────────

    def _apply_style(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure(".", background=self.BG, foreground=self.FG, font=self.FONT)
        style.configure("TFrame", background=self.BG)
        style.configure("TLabel", background=self.BG, foreground=self.FG, font=self.FONT)
        style.configure("TLabelframe", background=self.BG, foreground=self.ACCENT, font=self.FONT_BOLD)
        style.configure("TLabelframe.Label", background=self.BG, foreground=self.ACCENT, font=self.FONT_BOLD)
        style.configure("TEntry", fieldbackground=self.SURFACE, foreground=self.FG, insertcolor=self.FG)
        style.configure("TCombobox", fieldbackground=self.SURFACE, foreground=self.FG,
                         selectbackground=self.ACCENT, selectforeground=self.BG)

        # Buttons
        style.configure("Accent.TButton", background=self.ACCENT, foreground=self.BG,
                         font=self.FONT_BOLD, padding=(12, 6))
        style.map("Accent.TButton",
                   background=[("active", self.ACCENT_HOVER), ("disabled", self.SURFACE2)],
                   foreground=[("disabled", self.SURFACE)])

        style.configure("Export.TButton", background=self.GREEN, foreground=self.BG,
                         font=self.FONT_BOLD, padding=(12, 6))
        style.map("Export.TButton",
                   background=[("active", "#b5f0b0"), ("disabled", self.SURFACE2)],
                   foreground=[("disabled", self.SURFACE)])

        # Progressbar
        style.configure("Accent.Horizontal.TProgressbar",
                         troughcolor=self.SURFACE, background=self.ACCENT, thickness=8)

        # Treeview
        style.configure("Treeview",
                         background=self.SURFACE, foreground=self.FG, fieldbackground=self.SURFACE,
                         font=self.FONT, rowheight=26, borderwidth=0)
        style.configure("Treeview.Heading",
                         background=self.SURFACE2, foreground=self.ACCENT,
                         font=self.FONT_BOLD, borderwidth=0)
        style.map("Treeview",
                   background=[("selected", self.ACCENT)],
                   foreground=[("selected", self.BG)])

    # ── Build UI ────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Title ───────────────────────────────────────────────────────
        title_frame = ttk.Frame(self.root)
        title_frame.pack(fill="x", padx=20, pady=(15, 5))
        ttk.Label(title_frame, text="🌐 Fan-Out Lite Generator",
                  font=self.FONT_TITLE, foreground=self.ACCENT).pack(anchor="w")
        ttk.Label(title_frame, text="Top 5 queries sémantiques par mot-clé — FR · EN · DE · ES · PT-BR",
                  foreground=self.SURFACE2).pack(anchor="w")

        # ── Top controls frame ──────────────────────────────────────────
        ctrl_frame = ttk.Frame(self.root)
        ctrl_frame.pack(fill="x", padx=20, pady=(10, 5))

        # API Key
        key_frame = ttk.LabelFrame(ctrl_frame, text="  🔑 Clé API OpenAI  ", padding=8)
        key_frame.pack(side="left", fill="x", expand=True, padx=(0, 10))

        self.api_key_var = tk.StringVar()
        self.api_key_entry = ttk.Entry(key_frame, textvariable=self.api_key_var, show="*", width=50)
        self.api_key_entry.pack(fill="x")

        self.key_status = ttk.Label(key_frame, text="", foreground=self.SURFACE2)
        self.key_status.pack(anchor="w")

        # Language
        lang_frame = ttk.LabelFrame(ctrl_frame, text="  🌍 Langue  ", padding=8)
        lang_frame.pack(side="left", padx=(0, 10))

        self.lang_var = tk.StringVar(value="FR - Français")
        lang_combo = ttk.Combobox(lang_frame, textvariable=self.lang_var,
                                   values=list(LANGUAGES.keys()), state="readonly", width=20)
        lang_combo.pack()

        # Parallel workers
        para_frame = ttk.LabelFrame(ctrl_frame, text="  ⚡ Threads  ", padding=8)
        para_frame.pack(side="left")

        self.workers_var = tk.IntVar(value=5)
        workers_spin = ttk.Spinbox(para_frame, from_=1, to=10, textvariable=self.workers_var,
                                    width=4, state="readonly")
        workers_spin.pack()

        # ── Keywords input ──────────────────────────────────────────────
        kw_frame = ttk.LabelFrame(self.root, text="  📝 Mots-clés (un par ligne)  ", padding=8)
        kw_frame.pack(fill="x", padx=20, pady=(10, 5))

        self.kw_text = tk.Text(kw_frame, height=6, bg=self.SURFACE, fg=self.FG,
                                insertbackground=self.FG, font=self.FONT_MONO,
                                relief="flat", padx=8, pady=6, wrap="word",
                                selectbackground=self.ACCENT, selectforeground=self.BG)
        self.kw_text.pack(fill="x")

        # ── Action bar ──────────────────────────────────────────────────
        action_frame = ttk.Frame(self.root)
        action_frame.pack(fill="x", padx=20, pady=(5, 5))

        self.gen_btn = ttk.Button(action_frame, text="🚀  Générer le Fan-Out",
                                   style="Accent.TButton", command=self._on_generate)
        self.gen_btn.pack(side="left")

        self.export_btn = ttk.Button(action_frame, text="📥  Exporter XLSX",
                                      style="Export.TButton", command=self._on_export,
                                      state="disabled")
        self.export_btn.pack(side="left", padx=(10, 0))

        # Progress
        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(action_frame, variable=self.progress_var,
                                             maximum=100, style="Accent.Horizontal.TProgressbar",
                                             length=200)
        self.progress_bar.pack(side="left", padx=(15, 5))

        self.status_var = tk.StringVar(value="")
        ttk.Label(action_frame, textvariable=self.status_var,
                  foreground=self.YELLOW, font=self.FONT).pack(side="left", padx=5)

        # ── Results Treeview ────────────────────────────────────────────
        result_frame = ttk.LabelFrame(self.root, text="  📊 Résultats  ", padding=8)
        result_frame.pack(fill="both", expand=True, padx=20, pady=(5, 15))

        columns = ("keyword", "topic", "rank", "query")
        self.tree = ttk.Treeview(result_frame, columns=columns, show="headings", selectmode="extended")
        self.tree.heading("keyword", text="Mot-clé")
        self.tree.heading("topic", text="Topic")
        self.tree.heading("rank", text="#")
        self.tree.heading("query", text="Query Fan-Out")

        self.tree.column("keyword", width=180, minwidth=120)
        self.tree.column("topic", width=180, minwidth=100)
        self.tree.column("rank", width=40, minwidth=30, anchor="center")
        self.tree.column("query", width=450, minwidth=200)

        scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # ── Tag for alternating row colors ──────────────────────────────
        self.tree.tag_configure("even", background=self.SURFACE)
        self.tree.tag_configure("odd", background=self.SURFACE2)
        self.tree.tag_configure("error", background="#45273a", foreground=self.RED)

    # ── Load .env key ───────────────────────────────────────────────────

    def _load_env_key(self):
        """Pre-fill API key from .env / environment."""
        try:
            from dotenv import load_dotenv
            env_path = os.path.join(_PROJECT_ROOT, ".env")
            if os.path.exists(env_path):
                load_dotenv(env_path)
        except ImportError:
            pass

        key = os.getenv("OPENAI_API_KEY", "")
        if key:
            self.api_key_var.set(key)
            self.key_status.configure(text="✅ Clé chargée depuis .env", foreground=self.GREEN)
        else:
            self.key_status.configure(text="⚠ Saisissez votre clé OpenAI", foreground=self.YELLOW)

    # ── Parse keywords ──────────────────────────────────────────────────

    def _parse_keywords(self) -> List[str]:
        raw = self.kw_text.get("1.0", "end").strip()
        return [k.strip() for k in raw.splitlines() if k.strip()]

    # ── Generate (UI handler) ───────────────────────────────────────────

    def _on_generate(self):
        if self._generating:
            return

        api_key = self.api_key_var.get().strip()
        if not api_key:
            messagebox.showwarning("Clé API manquante",
                                   "Veuillez saisir votre clé API OpenAI.")
            return

        keywords = self._parse_keywords()
        if not keywords:
            messagebox.showwarning("Mots-clés manquants",
                                   "Veuillez saisir au moins un mot-clé.")
            return

        lang_code = LANGUAGES[self.lang_var.get()]

        # Set API key in env so OpenAIClient picks it up
        os.environ["OPENAI_API_KEY"] = api_key

        self._generating = True
        self._results.clear()
        self.tree.delete(*self.tree.get_children())
        self.gen_btn.configure(state="disabled")
        self.export_btn.configure(state="disabled")
        self.progress_var.set(0)
        self.status_var.set("Démarrage…")

        thread = threading.Thread(target=self._worker, args=(keywords, lang_code), daemon=True)
        thread.start()

    # ── Worker thread ───────────────────────────────────────────────────

    def _worker(self, keywords: List[str], lang_code: str):
        total = len(keywords)
        max_workers = min(self.workers_var.get(), total)
        completed = 0
        lock = threading.Lock()

        def _process_one(idx_kw):
            idx, kw = idx_kw
            gen = FanoutGenerator()  # one per thread for safety
            result = gen.generate(kw, language=lang_code)
            top_queries = FanoutGenerator.extract_top_queries(result, top_n=5)
            return idx, kw, result, top_queries

        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = {pool.submit(_process_one, (i, kw)): i
                       for i, kw in enumerate(keywords, 1)}

            for future in as_completed(futures):
                idx, kw, result, top_queries = future.result()
                entry = {
                    "keyword": kw,
                    "topic": result.topic or kw,
                    "queries": [(i, q) for i, q in enumerate(top_queries, 1)],
                    "error": result.error,
                }
                with lock:
                    self._results.append(entry)
                    completed += 1

                self.root.after(0, self._update_progress, completed, total, kw)
                self.root.after(0, self._insert_result_rows, entry, idx)

        self.root.after(0, self._on_complete, total)

    def _update_progress(self, current: int, total: int, kw: str):
        pct = current / total * 100
        self.progress_var.set(pct)
        self.status_var.set(f"{current}/{total} — {kw}")

    def _insert_result_rows(self, entry: dict, idx: int):
        tag = "even" if idx % 2 == 0 else "odd"

        if entry["error"] and not entry["queries"]:
            self.tree.insert("", "end",
                             values=(entry["keyword"], entry["topic"], "—", f"⚠ {entry['error']}"),
                             tags=("error",))
            return

        for rank, query in entry["queries"]:
            kw_display = entry["keyword"] if rank == 1 else ""
            topic_display = entry["topic"] if rank == 1 else ""
            self.tree.insert("", "end",
                             values=(kw_display, topic_display, rank, query),
                             tags=(tag,))

    def _on_complete(self, total: int):
        self.progress_var.set(100)
        n_ok = sum(1 for r in self._results if not r["error"])
        n_err = total - n_ok
        status = f"✅ {n_ok}/{total} mots-clés traités"
        if n_err:
            status += f" — {n_err} erreur(s)"
        self.status_var.set(status)
        self.gen_btn.configure(state="normal")
        if self._results:
            self.export_btn.configure(state="normal")
        self._generating = False

    # ── Export XLSX ─────────────────────────────────────────────────────

    def _on_export(self):
        if not self._results:
            messagebox.showinfo("Rien à exporter", "Lancez d'abord une génération.")
            return

        filepath = filedialog.asksaveasfilename(
            title="Enregistrer le fichier Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
            initialfile="fanout_lite_results.xlsx",
        )
        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Fan-Out Lite"

            # Header style
            header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                bottom=Side(style="thin", color="cccccc"),
            )

            headers = ["Mot-clé", "Topic",
                       "Fan-Out 1", "Fan-Out 2", "Fan-Out 3", "Fan-Out 4", "Fan-Out 5",
                       "Langue"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            lang_label = self.lang_var.get()
            row = 2
            for entry in self._results:
                ws.cell(row=row, column=1, value=entry["keyword"])
                ws.cell(row=row, column=2, value=entry["topic"])

                if entry["error"] and not entry["queries"]:
                    ws.cell(row=row, column=3, value=f"ERREUR: {entry['error']}")
                else:
                    for rank, query in entry["queries"]:
                        ws.cell(row=row, column=2 + rank, value=query)

                ws.cell(row=row, column=8, value=lang_label)
                for c in range(1, 9):
                    ws.cell(row=row, column=c).border = thin_border
                row += 1

            # Column widths
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 28
            for col_letter in ("C", "D", "E", "F", "G"):
                ws.column_dimensions[col_letter].width = 45
            ws.column_dimensions["H"].width = 18

            # Freeze header
            ws.freeze_panes = "A2"

            wb.save(filepath)
            messagebox.showinfo("Export réussi", f"Fichier enregistré :\n{filepath}")

        except ImportError:
            messagebox.showerror("Dépendance manquante",
                                 "Le package 'openpyxl' est requis.\n\npip install openpyxl")
        except (OSError, PermissionError) as e:
            messagebox.showerror("Erreur d'export", str(e))


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    FanoutLiteApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
