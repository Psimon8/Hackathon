"""
Unified Excel Exporter — multi-tab XLSX for all modules.
Combines data from SERP, Semantic Score, EEAT, Fan-out, Travel Agent.
"""
import io
import logging
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import (
    EEATResult,
    FanoutResult,
    KeywordVolumeResult,
    SemanticScoreResult,
    SERPResult,
)

logger = logging.getLogger(__name__)

# ── style constants ─────────────────────────────────────────────────────────
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _autofit(ws) -> None:
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 60)


# ═══════════════════════════════════════════════════════════════════════════
# Per-module sheet writers
# ═══════════════════════════════════════════════════════════════════════════

def _write_serp(wb: Workbook, results: List[SERPResult], keyword: str = "") -> None:
    ws = wb.create_sheet("SERP")
    headers = ["Position", "URL", "Domaine", "Titre", "Description", "Type"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in results:
        ws.append([r.position, r.url, r.domain, r.title, r.description, r.result_type])
    _autofit(ws)


def _write_semantic(wb: Workbook, results: List[SemanticScoreResult]) -> None:
    # Master sheet
    ws = wb.create_sheet("Semantic Score")
    headers = ["Keyword", "Avg Score", "Avg Competitor", "Domain URL", "Domain Score",
               "Domain Position", "Density (%)", "Time (s)", "Error"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in results:
        ws.append([
            r.keyword,
            round(r.average_score, 2) if r.average_score is not None else "",
            round(r.average_competitor_score, 2) if r.average_competitor_score is not None else "",
            r.domain_url or "",
            round(r.domain_score, 2) if r.domain_score is not None else "",
            r.domain_position if r.domain_position is not None else "",
            round(r.keyword_density, 3) if r.keyword_density else "",
            round(r.analysis_time, 1) if r.analysis_time else "",
            r.error or "",
        ])
    _autofit(ws)

    # Per-keyword detail sheets (max 10 to avoid Excel crash)
    for r in results[:10]:
        if not r.top_results:
            continue
        name = r.keyword[:28].replace("/", "-")  # sheet name limit
        ws2 = wb.create_sheet(f"S_{name}")
        h = ["Pos", "URL", "Titre", "Score", "Mots", "Méthode"]
        ws2.append(h)
        _style_header(ws2, len(h))
        for u in r.top_results:
            ws2.append([u.position, u.url, u.title, u.semantic_score, u.word_count, u.scrape_method])

        # ── N-grams Analysis ────────────────────────────────────────────
        cur_row = ws2.max_row + 3
        ws2.cell(row=cur_row, column=1, value="ANALYSE N-GRAMS").font = Font(bold=True, size=12)
        cur_row += 1

        for ng_type in ["unigrams", "bigrams", "trigrams"]:
            dom = r.domain_ngrams.get(ng_type, {}) if r.domain_ngrams else {}
            comp = r.average_competitor_ngrams.get(ng_type, {}) if r.average_competitor_ngrams else {}
            diff_map = (r.ngram_differential or {}).get(ng_type, {})
            all_terms = sorted(set(list(dom.keys()) + list(comp.keys())),
                               key=lambda t: diff_map.get(t, dom.get(t, 0) - comp.get(t, 0)))
            if not all_terms:
                continue

            ws2.cell(row=cur_row, column=1, value=ng_type.upper()).font = Font(bold=True)
            cur_row += 1
            ng_headers = ["N-gram", "Occ. Domaine", "Occ. Concurrent (moy.)", "Différence"]
            for ci, hdr in enumerate(ng_headers, 1):
                cell = ws2.cell(row=cur_row, column=ci, value=hdr)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
            cur_row += 1
            for term in all_terms:
                d_val = dom.get(term, 0)
                c_val = round(comp.get(term, 0), 1)
                diff_val = round(diff_map.get(term, d_val - c_val), 1)
                ws2.cell(row=cur_row, column=1, value=term)
                ws2.cell(row=cur_row, column=2, value=d_val)
                ws2.cell(row=cur_row, column=3, value=c_val)
                ws2.cell(row=cur_row, column=4, value=diff_val)
                cur_row += 1
            cur_row += 1  # blank row between ngram types

        # ── GPT Refined Occurrences ─────────────────────────────────────
        refined = getattr(r, 'refined_ngrams', None)
        if refined:
            cur_row = ws2.max_row + 3
            ws2.cell(row=cur_row, column=1, value="OCCURRENCES RAFFINÉES (GPT)").font = Font(bold=True, size=12)
            cur_row += 1
            ref_headers = ["N-gram", "Type", "Catégorie", "Priorité SEO", "Occ. Domaine", "Occ. Concurrent"]
            for ci, hdr in enumerate(ref_headers, 1):
                cell = ws2.cell(row=cur_row, column=ci, value=hdr)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
            cur_row += 1
            for ng in refined:
                ws2.cell(row=cur_row, column=1, value=ng.get("ngram", ""))
                ws2.cell(row=cur_row, column=2, value=ng.get("type", ""))
                ws2.cell(row=cur_row, column=3, value=ng.get("category", ""))
                ws2.cell(row=cur_row, column=4, value=ng.get("priority_score", 0))
                ws2.cell(row=cur_row, column=5, value=ng.get("occurrences_domain", 0))
                ws2.cell(row=cur_row, column=6, value=ng.get("occurrences_competitor", 0))
                cur_row += 1

        # ── SEO Brief ───────────────────────────────────────────────────
        seo_brief = getattr(r, 'seo_brief', None)
        if seo_brief:
            brief = seo_brief
            cur_row = ws2.max_row + 3
            ws2.cell(row=cur_row, column=1, value="BRIEF SEO").font = Font(bold=True, size=12)
            cur_row += 1

            for label, key in [("Title", "title"), ("Meta Description", "meta_description"), ("H1", "h1")]:
                ws2.cell(row=cur_row, column=1, value=label).font = Font(bold=True)
                ws2.cell(row=cur_row, column=2, value=brief.get(key, ""))
                cur_row += 1

            if brief.get("target_word_count"):
                ws2.cell(row=cur_row, column=1, value="Nombre de mots cible").font = Font(bold=True)
                ws2.cell(row=cur_row, column=2, value=brief["target_word_count"])
                cur_row += 1

            sections = brief.get("sections", [])
            if sections:
                cur_row += 1
                ws2.cell(row=cur_row, column=1, value="STRUCTURE Hn RECOMMANDÉE").font = Font(bold=True)
                cur_row += 1
                sec_headers = ["Niveau", "Heading", "Description contenu"]
                for ci, hdr in enumerate(sec_headers, 1):
                    cell = ws2.cell(row=cur_row, column=ci, value=hdr)
                    cell.font = _HEADER_FONT
                    cell.fill = _HEADER_FILL
                    cell.alignment = Alignment(horizontal="center")
                cur_row += 1
                for sec in sections:
                    ws2.cell(row=cur_row, column=1, value=sec.get("level", "h2").upper())
                    ws2.cell(row=cur_row, column=2, value=sec.get("heading", ""))
                    ws2.cell(row=cur_row, column=3, value=sec.get("content_description", ""))
                    cur_row += 1

        _autofit(ws2)


def _write_eeat(wb: Workbook, results: List[EEATResult]) -> None:
    ws = wb.create_sheet("EEAT Scoring")
    headers = [
        "URL", "Titre", "Langue", "EEAT Global", "Expertise", "Experience",
        "Authoritativeness", "Trustworthiness", "Sentiment", "Lisibilité",
        "Catégorie", "Composite", "Compliance", "Qualité", "Entité", "Résumé", "Statut",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in results:
        comp = r.eeat_components or {}
        ws.append([
            r.url, r.title, r.language, r.eeat_global,
            comp.get("expertise", ""), comp.get("experience", ""),
            comp.get("authoritativeness", ""), comp.get("trustworthiness", ""),
            r.sentiment, r.lisibilite_score, r.categorie,
            r.composite_score, r.compliance_score, r.quality_level,
            r.main_entity, r.resume, r.status,
        ])

    # Suggestions sub-sheet
    ws2 = wb.create_sheet("EEAT Suggestions")
    ws2.append(["URL", "Suggestion"])
    _style_header(ws2, 2)
    for r in results:
        for s in (r.suggestions or []):
            ws2.append([r.url, s])
    _autofit(ws)
    _autofit(ws2)


def _write_fanout(wb: Workbook, results: List[FanoutResult]) -> None:
    ws = wb.create_sheet("Fan-out")
    headers = ["Keyword", "Topic", "Top 3 Questions", "Mandatory Queries", "Recommended Queries", "Optional Queries", "Justification"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in results:
        mq = "; ".join(q for f in r.mandatory for q in f.queries)
        rq = "; ".join(q for f in r.recommended for q in f.queries)
        oq = "; ".join(q for f in r.optional for q in f.queries)
        ws.append([
            r.keyword, r.topic,
            " | ".join(r.top_3_questions),
            mq, rq, oq, r.justification,
        ])
    _autofit(ws)


def _write_volumes(wb: Workbook, results: List[KeywordVolumeResult]) -> None:
    ws = wb.create_sheet("Search Volumes")
    headers = ["Keyword", "Volume", "Competition", "CPC", "Origin"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in results:
        ws.append([
            r.keyword, r.search_volume, r.competition, r.cpc,
            r.origin or "",
        ])
    _autofit(ws)


# ═══════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════

def export_to_excel(
    *,
    serp_results: Optional[List[SERPResult]] = None,
    semantic_results: Optional[List[SemanticScoreResult]] = None,
    eeat_results: Optional[List[EEATResult]] = None,
    fanout_results: Optional[List[FanoutResult]] = None,
    volume_results: Optional[List[KeywordVolumeResult]] = None,
    filename: Optional[str] = None,
) -> bytes:
    """
    Build a multi-tab XLSX workbook and return raw bytes.
    If *filename* is given, also write to disk.
    """
    wb = Workbook()
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    if serp_results:
        _write_serp(wb, serp_results)
    if semantic_results:
        _write_semantic(wb, semantic_results)
    if eeat_results:
        _write_eeat(wb, eeat_results)
    if fanout_results:
        _write_fanout(wb, fanout_results)
    if volume_results:
        _write_volumes(wb, volume_results)

    if not wb.sheetnames:
        ws = wb.create_sheet("Info")
        ws.append(["Aucune donnée à exporter."])

    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    if filename:
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        with open(filename, "wb") as f:
            f.write(raw)
        logger.info("Exported %s (%d KB)", filename, len(raw) // 1024)

    return raw


def default_filename(prefix: str = "yn_export") -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.xlsx"
